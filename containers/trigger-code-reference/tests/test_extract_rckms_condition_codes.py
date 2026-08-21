import csv

import pytest
from openpyxl import Workbook

from data.extract_rckms_condition_codes import (
    EXPECTED_HEADERS,
    SHEET_NAME,
    _extract_condition_codes,
)


def test_extracts_named_sheet_and_normalizes_headers(tmp_path):
    source = tmp_path / "RCKMS Condition Codes_20260626.xlsx"
    destination = tmp_path / "RCKMS Condition Codes.csv"

    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read Me"
    worksheet = workbook.create_sheet(SHEET_NAME)
    worksheet.append(["Condition\n", *EXPECTED_HEADERS[1:]])
    worksheet.append(
        [
            "Acanthamoeba",
            "Waterborne (not enteric)",
            49649001,
            "Infection caused by Acanthamoeba (disorder)",
            "SNOMEDCT",
            "2026-03",
            None,
            "50225  Acanthamoeba disease (excluding keratitis)",
            "Previously released condition",
        ]
    )
    workbook.save(source)

    row_count = _extract_condition_codes(source, destination)

    assert row_count == 1
    with destination.open(encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.reader(csv_file))
    assert rows[0] == list(EXPECTED_HEADERS)
    assert rows[1][2] == "49649001"


def test_rejects_an_unexpected_header(tmp_path):
    source = tmp_path / "RCKMS Condition Codes.xlsx"
    destination = tmp_path / "RCKMS Condition Codes.csv"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(["Unexpected", *EXPECTED_HEADERS[1:]])
    worksheet.append(["value"] * len(EXPECTED_HEADERS))
    workbook.save(source)

    with pytest.raises(ValueError, match="Unexpected RCKMS header row"):
        _extract_condition_codes(source, destination)
