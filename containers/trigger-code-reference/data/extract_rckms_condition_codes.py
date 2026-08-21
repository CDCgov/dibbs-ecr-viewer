import argparse
import csv
import os
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

SHEET_NAME = "RCKMS Condition Codes"
# These columns are the contract between the downloaded workbook and the seed
# script. Failing when they change is safer than silently building a database
# from a worksheet whose layout we no longer understand.
EXPECTED_HEADERS = (
    "Condition",
    "Condition Category",
    "Code",
    "Code Description",
    "Code System",
    "Code System Version",
    "Planned Update",
    "NNC Code",
    "Content Release",
)


def _normalize_header(value: object) -> str:
    # The workbook currently contains a line break in the first header cell.
    # Header whitespace does not carry meaning, so remove it before validation.
    return "" if value is None else str(value).strip()


def _serialize_value(value: object) -> str:
    # Empty spreadsheet cells should become empty CSV fields.
    if value is None:
        return ""
    # Excel may return an identifier such as a SNOMED code as 49649001.0.
    # Write whole numbers without the decimal so they remain valid codes.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _extract_condition_codes(source: Path, destination: Path) -> int:
    # Read values only. The condition-code worksheet contains source data, not
    # formulas that need to be copied into the generated CSV.
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(
                f'Workbook is missing the "{SHEET_NAME}" sheet. '
                f"Available sheets: {available_sheets}"
            )

        worksheet = workbook[SHEET_NAME]
        rows = worksheet.iter_rows(values_only=True)

        try:
            first_row = next(rows)
        except StopIteration as error:
            raise ValueError(f'The "{SHEET_NAME}" sheet is empty.') from error

        actual_headers = tuple(
            _normalize_header(value) for value in first_row[: len(EXPECTED_HEADERS)]
        )
        extra_headers = first_row[len(EXPECTED_HEADERS) :]
        if actual_headers != EXPECTED_HEADERS or any(extra_headers):
            raise ValueError(
                "Unexpected RCKMS header row.\n"
                f"Expected: {EXPECTED_HEADERS}\n"
                f"Actual:   {actual_headers}"
            )

        extracted_rows = []
        for row_number, row in enumerate(rows, start=2):
            # Columns after I are not part of the known file format. Reject
            # populated extra columns so an upstream format change is visible.
            values = row[: len(EXPECTED_HEADERS)]
            extra_values = row[len(EXPECTED_HEADERS) :]
            if any(value is not None for value in extra_values):
                raise ValueError(f"Unexpected data after column I on row {row_number}.")
            if all(value is None or str(value).strip() == "" for value in values):
                # Ignore visual spacer rows if RCKMS adds any to the worksheet.
                continue
            extracted_rows.append([_serialize_value(value) for value in values])
    finally:
        workbook.close()

    if not extracted_rows:
        raise ValueError(f'The "{SHEET_NAME}" sheet contains no condition rows.')

    destination.parent.mkdir(parents=True, exist_ok=True)
    # Write to a temporary file and replace the destination only after the
    # entire CSV succeeds. This avoids leaving a partial generated file behind.
    temporary_file = NamedTemporaryFile(
        "w",
        encoding="utf-8-sig",
        newline="",
        dir=destination.parent,
        delete=False,
    )
    try:
        with temporary_file:
            writer = csv.writer(temporary_file, lineterminator="\n")
            writer.writerow(EXPECTED_HEADERS)
            writer.writerows(extracted_rows)
        os.replace(temporary_file.name, destination)
    except BaseException:
        Path(temporary_file.name).unlink(missing_ok=True)
        raise

    return len(extracted_rows)


def _main() -> None:
    parser = argparse.ArgumentParser(
        description='Extract the "RCKMS Condition Codes" worksheet as CSV.'
    )
    parser.add_argument("source", type=Path, help="Downloaded RCKMS .xlsx file")
    parser.add_argument("destination", type=Path, help="Destination CSV file")
    args = parser.parse_args()

    row_count = _extract_condition_codes(args.source, args.destination)
    print(f"Extracted {row_count} condition rows to {args.destination}")


if __name__ == "__main__":
    _main()
